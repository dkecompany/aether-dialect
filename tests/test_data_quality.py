"""Tests for CSV/Excel upload validation."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pytest

from aetherdialect._constants_runtime import (
    DATA_QUALITY_DETAIL_CANDIDATE_HEADER_ROW,
    DATA_QUALITY_ISSUE_APPEND_HEADER_MISMATCH,
    DATA_QUALITY_ISSUE_APPENDABLE_REGIONS,
    DATA_QUALITY_ISSUE_DUPLICATE_HEADER,
    DATA_QUALITY_ISSUE_EXCEL_ERROR,
    DATA_QUALITY_ISSUE_HEADER_NOT_ROW_ONE,
    DATA_QUALITY_ISSUE_MIXED_TYPES,
    DATA_QUALITY_ISSUE_NULL_TOKEN,
    DATA_QUALITY_ISSUE_NUMBER_AS_TEXT,
    DATA_QUALITY_ISSUE_REPEATED_HEADER,
    DATA_QUALITY_ISSUE_SECTION_HEADING,
)
from aetherdialect._contracts_schema import CsvSourceSelection
from aetherdialect._data_quality import (
    apply_source_selection,
    detect_grid_issues,
    load_source_grids,
    normalize_grid,
    prepare_relations_for_paths,
    validate_upload_sources,
)


def _mock_llm_json(system: str, user: str, retries: int = 1, task: str = "default") -> dict[str, object]:
    if task == "upload_summary":
        return {"summary": "Upload inspection completed."}
    if task == "upload_interpret":
        return {}
    raise AssertionError(f"unexpected llm_json task={task!r}")


@pytest.fixture(autouse=True)
def _patch_upload_llm() -> None:
    with patch("aetherdialect._data_quality.LLMProvider.json", side_effect=_mock_llm_json):
        yield


@pytest.mark.fast
def test_detect_duplicate_header_includes_location(tmp_path: Path) -> None:
    path = tmp_path / "items.csv"
    path.write_text("id,id\n1,2\n", encoding="utf-8")
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    assert any(
        any(key == "issue_code" and value == DATA_QUALITY_ISSUE_DUPLICATE_HEADER for key, value in issue.details)
        for issue in issues
    )
    assert any(any(key == "location" and value.endswith("!B1") for key, value in issue.details) for issue in issues)


@pytest.mark.fast
def test_validate_clean_csv_returns_ok(tmp_path: Path) -> None:
    path = tmp_path / "customers.csv"
    path.write_text("id,name\n1,Alice\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok


@pytest.mark.fast
def test_validate_ragged_row_ok_after_reshape(tmp_path: Path) -> None:
    path = tmp_path / "orders.csv"
    path.write_text("id,amount\n10,25.5\n99\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok


@pytest.mark.fast
def test_embedded_newline_preserved_in_cell(tmp_path: Path) -> None:
    path = tmp_path / "survey.csv"
    path.write_text('question\n"line one\nline two"\n', encoding="utf-8")
    grid = normalize_grid(load_source_grids(path)[0], normalize_cell_newlines=True)
    assert any("line one\nline two" in cell for row in grid.cells for cell in row)


@pytest.mark.fast
def test_header_not_row_one_is_advisory(tmp_path: Path) -> None:
    path = tmp_path / "shifted.csv"
    path.write_text("Survey export\nid,name\n1,Alice\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.requires_review
    assert any(
        any(key == "issue_code" and value == DATA_QUALITY_ISSUE_HEADER_NOT_ROW_ONE for key, value in issue.details)
        for issue in report.issues
    )


@pytest.mark.fast
def test_source_selection_skips_metadata_rows(tmp_path: Path) -> None:
    path = tmp_path / "shifted.csv"
    path.write_text("Survey export\nid,name\n1,Alice\n", encoding="utf-8")
    relations = prepare_relations_for_paths(
        (path,),
        source_selections={path.name: CsvSourceSelection(header_row=2)},
    )
    assert len(relations) == 1
    assert relations[0].columns == ("id", "name")


@pytest.mark.fast
def test_semicolon_delimited_csv_loads(tmp_path: Path) -> None:
    path = tmp_path / "semi.csv"
    path.write_text("id;name\n1;Alice\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok


@pytest.mark.fast
def test_detect_overfull_row(tmp_path: Path) -> None:
    path = tmp_path / "overfull.csv"
    path.write_text("id,amount\n1,10,extra\n", encoding="utf-8")
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    assert any(any(key == "issue_code" and value == "overfull_row" for key, value in issue.details) for issue in issues)


@pytest.mark.fast
def test_single_column_csv_warns_but_loads(tmp_path: Path) -> None:
    path = tmp_path / "single.csv"
    path.write_text("only\nvalue\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok
    relations = prepare_relations_for_paths((path,))
    assert len(relations) == 1


@pytest.mark.fast
def test_stacked_workbook_ok_after_reshape(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "stacked.xlsx"
    wb = Workbook()
    ws = wb.active
    header = ["Metric", "Q1", "Q2"]
    ws.append(header)
    ws.append(["Revenue", 100, 110])
    ws.append(header)
    ws.append(["Costs", 50, 60])
    wb.save(path)
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok


@pytest.mark.fast
def test_detect_repeated_header_in_stacked_workbook(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "stacked.xlsx"
    wb = Workbook()
    ws = wb.active
    header = ["Metric", "Q1", "Q2"]
    ws.append(header)
    ws.append(["Revenue", 100, 110])
    ws.append(header)
    ws.append(["Costs", 50, 60])
    wb.save(path)
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    assert any(
        any(key == "issue_code" and value == DATA_QUALITY_ISSUE_REPEATED_HEADER for key, value in issue.details)
        for issue in issues
    )


@pytest.mark.fast
def test_section_heading_splits_into_multiple_relations(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Metric", "Q1", "Q2"])
    ws.append(["Revenue", 100, 110])
    ws.append(["Section B Division", None, None])
    ws.append(["Costs", 50, 60])
    ws.append(["Section C Division", None, None])
    wb.save(path)
    relations = prepare_relations_for_paths((path,))
    assert len(relations) >= 2


@pytest.mark.fast
def test_blank_separated_regions_validate_ok(tmp_path: Path) -> None:
    path = tmp_path / "multi.csv"
    path.write_text("id,name\n1,Alice\n\nid,name\n2,Bob\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok


@pytest.mark.fast
def test_blank_separated_regions_produce_two_relations(tmp_path: Path) -> None:
    path = tmp_path / "multi.csv"
    path.write_text("id,name\n1,Alice\n\nid,name\n2,Bob\n", encoding="utf-8")
    relations = prepare_relations_for_paths((path,))
    assert len(relations) == 2
    names = {relation.relation_name for relation in relations}
    assert names == {"multi__1", "multi__2"}
    assert relations[0].rows[0]["id"] == "1"
    assert relations[1].rows[0]["id"] == "2"


@pytest.mark.fast
def test_duplicate_header_message_includes_column_location(tmp_path: Path) -> None:
    path = tmp_path / "items.csv"
    path.write_text("id,id\n1,2\n", encoding="utf-8")
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    duplicate = next(
        issue
        for issue in issues
        if any(key == "issue_code" and value == DATA_QUALITY_ISSUE_DUPLICATE_HEADER for key, value in issue.details)
    )
    assert "B1" in duplicate.message or any(key == "location" and "B1" in value for key, value in duplicate.details)


@pytest.mark.fast
def test_resolve_identifier_skips_llm_for_clean_names(monkeypatch: pytest.MonkeyPatch) -> None:
    from aetherdialect._data_quality import resolve_identifier_name

    def _fail_llm(*_args: object, **_kwargs: object) -> dict[str, str]:
        raise AssertionError("llm_json should not be called for clean identifiers")

    monkeypatch.setattr("aetherdialect._data_quality.LLMProvider.json", _fail_llm)
    reserved: set[str] = set()
    assert resolve_identifier_name("amount", kind="column", pinned_names={}, reserved=reserved) == "amount"
    assert resolve_identifier_name("order_date", kind="column", pinned_names={}, reserved=reserved) == "order_date"


@pytest.mark.fast
def test_data_quality_report_to_json_dict_shape(tmp_path: Path) -> None:
    path = tmp_path / "items.csv"
    path.write_text("id,name\n1,Alice\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    payload = report.to_json_dict()
    assert payload["ok"] is True
    assert isinstance(payload["narrative"], str)
    assert isinstance(payload["issues"], list)
    assert isinstance(payload["suggested_selections"], dict)
    if payload["issues"]:
        issue = payload["issues"][0]
        assert {"code", "level", "blocking", "location", "message", "details"} <= set(issue)


@pytest.mark.fast
def test_column_issue_precedence_collapses_to_excel_error(tmp_path: Path) -> None:
    path = tmp_path / "mixed.csv"
    path.write_text("amount,status\n#REF!,open\n12,closed\n", encoding="utf-8")
    grid = apply_source_selection(
        normalize_grid(load_source_grids(path)[0]),
        CsvSourceSelection(header_row=1),
    )
    issues = detect_grid_issues(grid)
    codes = {
        detail_value for issue in issues for detail_key, detail_value in issue.details if detail_key == "issue_code"
    }
    assert DATA_QUALITY_ISSUE_EXCEL_ERROR in codes
    assert DATA_QUALITY_ISSUE_MIXED_TYPES not in codes
    assert DATA_QUALITY_ISSUE_NULL_TOKEN not in codes


@pytest.mark.fast
def test_report_layout_suppresses_row_noise(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Metric", "Q1", "Q2"])
    ws.append(["Revenue", 100, 110])
    ws.append(["Section B Division", None, None])
    ws.append(["Costs", 50, 60])
    ws.append(["Section C Division", None, None])
    wb.save(path)
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    codes = {
        detail_value for issue in issues for detail_key, detail_value in issue.details if detail_key == "issue_code"
    }
    assert DATA_QUALITY_ISSUE_SECTION_HEADING in codes or len(codes) == 0
    assert "blank_row" not in codes
    assert "ragged_row" not in codes


@pytest.mark.fast
def test_appendable_regions_advisory_emitted(tmp_path: Path) -> None:
    path = tmp_path / "multi.csv"
    path.write_text("id,name\n1,Alice\n\nid,name\n2,Bob\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok
    assert any(
        any(key == "issue_code" and value == DATA_QUALITY_ISSUE_APPENDABLE_REGIONS for key, value in issue.details)
        for issue in report.issues
    )
    relations = prepare_relations_for_paths((path,))
    assert len(relations) == 2


@pytest.mark.fast
def test_append_regions_unions_into_one_relation(tmp_path: Path) -> None:
    path = tmp_path / "multi.csv"
    path.write_text("id,name\n1,Alice\n\nid,name\n2,Bob\n", encoding="utf-8")
    selection = CsvSourceSelection(append_regions=("A1:B2", "A4:B5"))
    relations = prepare_relations_for_paths((path,), source_selections={path.name: selection})
    assert len(relations) == 1
    assert len(relations[0].rows) == 2


@pytest.mark.fast
def test_append_regions_blocks_on_header_mismatch(tmp_path: Path) -> None:
    path = tmp_path / "multi.csv"
    path.write_text("id,name\n1,Alice\n\nsku,qty\n2,3\n", encoding="utf-8")
    selection = CsvSourceSelection(append_regions=("A1:B2", "A4:B5"))
    report = validate_upload_sources((path,), log_sink=lambda _msg: None, source_selections={path.name: selection})
    assert not report.ok
    assert any(
        any(key == "issue_code" and value == DATA_QUALITY_ISSUE_APPEND_HEADER_MISMATCH for key, value in issue.details)
        for issue in report.issues
    )


@pytest.mark.fast
def test_categorical_percent_not_number_as_text(tmp_path: Path) -> None:
    path = tmp_path / "buckets.csv"
    path.write_text("band,value\n26% - 50%,1\n10% - 25%,2\n", encoding="utf-8")
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    codes = {
        detail_value for issue in issues for detail_key, detail_value in issue.details if detail_key == "issue_code"
    }
    assert DATA_QUALITY_ISSUE_NUMBER_AS_TEXT not in codes


@pytest.mark.fast
def test_suggested_selections_shape(tmp_path: Path) -> None:
    path = tmp_path / "shifted.csv"
    path.write_text("Title\nid,name\n1,Alice\n", encoding="utf-8")
    report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert path.name in report.suggested_selections or report.suggested_selections == {}


@pytest.mark.fast
def test_llm_interpret_bad_proposal_rejected(tmp_path: Path) -> None:
    path = tmp_path / "ambig.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")

    def _bad_interpret(system: str, user: str, retries: int = 1, task: str = "default") -> dict[str, object]:
        if task == "upload_summary":
            return {"summary": "ok"}
        if task == "upload_interpret":
            return {"header_row": 999}
        if task == "upload_column_transforms":
            return {"column_transforms": []}
        return {"column_transforms": []}

    with patch("aetherdialect._data_quality.LLMProvider.json", side_effect=_bad_interpret):
        report = validate_upload_sources((path,), log_sink=lambda _msg: None)
    assert report.ok


@pytest.mark.fast
def test_transposed_key_value_orientation(tmp_path: Path) -> None:
    path = tmp_path / "kv.csv"
    path.write_text("field,value\nname,Alice\nage,30\n", encoding="utf-8")
    relations = prepare_relations_for_paths((path,))
    assert len(relations) == 1
    assert len(relations[0].rows) >= 1


@pytest.mark.fast
def test_header_not_row_one_includes_candidate(tmp_path: Path) -> None:
    path = tmp_path / "shifted.csv"
    path.write_text("Survey export\nid,name\n1,Alice\n", encoding="utf-8")
    grid = normalize_grid(load_source_grids(path)[0])
    issues = detect_grid_issues(grid)
    assert any(
        any(key == "issue_code" and value == DATA_QUALITY_ISSUE_HEADER_NOT_ROW_ONE for key, value in issue.details)
        for issue in issues
    )
    assert any(
        any(key == DATA_QUALITY_DETAIL_CANDIDATE_HEADER_ROW for key, _value in issue.details) for issue in issues
    )
